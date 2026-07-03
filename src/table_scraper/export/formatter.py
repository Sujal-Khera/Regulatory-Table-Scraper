"""Excel presentation formatting — freeze panes, bold headers, column widths, autofilters, conditional formatting."""

from __future__ import annotations

from typing import Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter


# Confidence conditional formatting fills
_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Charge value keywords for number formatting detection
_CHARGE_KEYWORDS = {"charge", "surcharge", "rs/kwh", "rs/kw", "value", "cost", "rate"}


def apply_workbook_formatting(workbook_path: str, format_config: dict[str, Any]) -> None:
    """Apply presentation formatting (fonts, widths, frozen header) to an Excel workbook.

    Modifies the workbook binary directly on disk.

    Args:
        workbook_path: File system path to the written .xlsx file.
        format_config: Presentation configuration parameters.
    """
    wb = openpyxl.load_workbook(workbook_path)

    # Resolve options from format config
    font_name = format_config.get("font_name", "Calibri")
    font_size = format_config.get("font_size", 11)
    freeze_header = format_config.get("freeze_header", True)
    wrap_text = format_config.get("wrap_text", True)
    auto_width = format_config.get("auto_width", True)

    header_font = Font(name=font_name, size=font_size, bold=True)
    body_font = Font(name=font_name, size=font_size, bold=False)
    align_wrap = Alignment(wrap_text=wrap_text, vertical="top")

    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 1:
            continue

        # 1. Freeze header row (A2)
        if freeze_header:
            ws.freeze_panes = "A2"

        # 2. Add autofilter on header row
        if ws.max_column and ws.max_column > 0 and ws.max_row > 1:
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        # 3. Identify confidence and charge columns by header name
        confidence_col = None
        charge_cols: list[int] = []
        header_values: dict[int, str] = {}

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            header_val = str(cell.value).strip().lower() if cell.value else ""
            header_values[col_idx] = header_val
            if header_val in ("confidence",):
                confidence_col = col_idx
            # Detect charge columns by keyword match
            if any(kw in header_val for kw in _CHARGE_KEYWORDS):
                charge_cols.append(col_idx)

        # 4. Bold header row (row 1) and align wrap
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = align_wrap

        # 5. Format body rows with conditional formatting and number formatting
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                if wrap_text:
                    cell.alignment = align_wrap

                # Confidence conditional formatting: green ≥0.9, yellow 0.7-0.9, red <0.7
                if col_idx == confidence_col and cell.value is not None:
                    try:
                        conf_val = float(cell.value)
                        if conf_val >= 0.9:
                            cell.fill = _FILL_GREEN
                        elif conf_val >= 0.7:
                            cell.fill = _FILL_YELLOW
                        else:
                            cell.fill = _FILL_RED
                    except (ValueError, TypeError):
                        pass

                # Number formatting for charge value columns (2 decimal places)
                if col_idx in charge_cols and cell.value is not None:
                    try:
                        float(cell.value)
                        cell.number_format = "0.00"
                    except (ValueError, TypeError):
                        pass

        # 6. Auto-fit column widths with bounds limits [10, 50]
        if auto_width:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        # Clean cell string representation
                        val_str = str(cell.value)
                        max_len = max(max_len, len(val_str))
                # Add breathing room padding
                ws.column_dimensions[col_letter].width = max(10, min(50, max_len + 3))

    wb.save(workbook_path)
    wb.close()
